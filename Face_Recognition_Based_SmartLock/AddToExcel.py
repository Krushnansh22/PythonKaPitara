# Tu nauhi bachega saleeee
def unauthorized_unlock(xlsx_file_path, image):
    import openpyxl
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, Alignment
    from datetime import datetime
    import os
    import cv2

    # Ensure the directory exists
    os.makedirs('UnauthorisedUnlocks', exist_ok=True)

    current_date = datetime.now().strftime('%d-%m-%Y')
    current_time = datetime.now().strftime('%H-%M-%S')

    # Save the image
    image_path = f'UnauthorisedUnlocks/{current_date}_{current_time}.jpg'
    cv2.imwrite(image_path, image)

    # Check if the file exists
    if not os.path.exists(xlsx_file_path):
        # Create a new workbook and add the headers
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        headers = ["Date", "Time", "Image"]

        # Set header style: bold, red color
        header_font = Font(bold=True, color="FF0000")

        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        workbook.save(xlsx_file_path)
    else:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(xlsx_file_path)
        sheet = workbook.active

    # Find the next available row
    next_row = sheet.max_row + 1

    # Define center alignment
    center_alignment = Alignment(horizontal='center')

    # Add date and time to the next row with center alignment
    date_cell = sheet.cell(row=next_row, column=1, value=current_date)
    date_cell.alignment = center_alignment

    time_cell = sheet.cell(row=next_row, column=2, value=current_time)
    time_cell.alignment = center_alignment

    # Set the row height to 100
    sheet.row_dimensions[next_row].height = 100

    # Set the column widths to 40
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20

    # Add the image to the next row
    img = Image(image_path)
    img.width, img.height = 100, 100  # Resize image to 100x100 if needed
    sheet.add_image(img, f'C{next_row}')

    # Save the workbook
    workbook.save(xlsx_file_path)

    print(f"Data added to row {next_row} in the Excel file at {xlsx_file_path}.")
